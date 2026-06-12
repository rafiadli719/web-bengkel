import openpyxl

file_path = r"C:\xampp\htdocs\web-bengkel\aplikasi\aplikasi\RENCANA ORDER 21052025.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    for sheet_name in wb.sheetnames:
        print(f"--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        # Get up to 20 rows
        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=False)):
            row_vals = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_vals.append("")
                elif isinstance(val, str) and val.startswith('='):
                    row_vals.append(f"FORMULA:{val}")
                else:
                    row_vals.append(str(val))
            print(" | ".join(row_vals))
            
except Exception as e:
    print(f"Error: {str(e)}")
