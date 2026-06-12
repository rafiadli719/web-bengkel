import openpyxl

file_path = r"C:\xampp\htdocs\web-bengkel\aplikasi\aplikasi\order.xlsx"
out_file = r"C:\xampp\htdocs\web-bengkel\aplikasi\aplikasi\excel_analysis.txt"

try:
    print(f"Opening {file_path}...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    
    with open(out_file, "w", encoding="utf-8") as f:
        for sheet_name in wb.sheetnames:
            f.write(f"\n=== Sheet: {sheet_name} ===\n")
            sheet = wb[sheet_name]
            
            # Find first row with data
            data_found = False
            for r_idx, row in enumerate(sheet.iter_rows()):
                row_vals = [cell.value for cell in row]
                if any(v is not None for v in row_vals):
                    f.write(f"Data starts at row {r_idx + 1}\n")
                    # Print this row and the next 100 rows, up to column 60 (column BH)
                    for i in range(r_idx, min(r_idx + 100, sheet.max_row if sheet.max_row else r_idx + 100)):
                        curr_row = list(sheet.iter_rows(min_row=i+1, max_row=i+1, max_col=60))[0]
                        vals = []
                        for cell in curr_row:
                            val = cell.value
                            if val is None:
                                vals.append("")
                            elif isinstance(val, str) and val.startswith('='):
                                vals.append(f"F:{val}")
                            else:
                                vals.append(str(val))
                        # Only print if row is not entirely empty
                        if any(v != "" for v in vals):
                            f.write(f"R{i+1}: " + " | ".join(vals) + "\n")
                    data_found = True
                    break
            if not data_found:
                f.write("No data found in this sheet.\n")
    print(f"Analysis saved to {out_file}")
            
except Exception as e:
    print(f"Error: {str(e)}")
