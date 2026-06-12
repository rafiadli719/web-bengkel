import openpyxl

file_path = r"C:\xampp\htdocs\web-bengkel\aplikasi\aplikasi\order.xlsx"

try:
    print(f"Opening {file_path}...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    
    for sheet_name in wb.sheetnames:
        print(f"\n=== Sheet: {sheet_name} ===")
        sheet = wb[sheet_name]
        
        # Find first row with data
        data_found = False
        for r_idx, row in enumerate(sheet.iter_rows()):
            row_vals = [cell.value for cell in row]
            if any(v is not None for v in row_vals):
                print(f"Data starts at row {r_idx + 1}")
                # Print this row and the next 50
                for i in range(r_idx, min(r_idx + 50, sheet.max_row if sheet.max_row else r_idx + 50)):
                    curr_row = list(sheet.iter_rows(min_row=i+1, max_row=i+1))[0]
                    vals = []
                    for cell in curr_row:
                        val = cell.value
                        if val is None:
                            vals.append("")
                        elif isinstance(val, str) and val.startswith('='):
                            vals.append(f"FORMULA:{val}")
                        else:
                            vals.append(str(val))
                    # Only print if row is not entirely empty
                    if any(v != "" for v in vals):
                        print(f"R{i+1}: " + " | ".join(vals))
                data_found = True
                break
        
        if not data_found:
            print("No data found in this sheet.")
            
except Exception as e:
    print(f"Error: {str(e)}")
