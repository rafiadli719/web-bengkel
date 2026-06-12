import openpyxl
import pandas as pd
import json

file_path = r"C:\xampp\htdocs\web-bengkel\aplikasi\aplikasi\RENCANA ORDER 21052025.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet_names = wb.sheetnames
    
    result = {
        "sheets": []
    }
    
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        data = []
        # Get first 50 rows for analysis
        for row in sheet.iter_rows(min_row=1, max_row=100, values_only=False):
            row_data = []
            for cell in row:
                row_data.append({
                    "value": str(cell.value) if cell.value is not None else "",
                    "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                })
            data.append(row_data)
        
        result["sheets"].append({
            "name": sheet_name,
            "data": data[:50] # Limit to first 50 rows for initial look
        })
    
    print(json.dumps(result, indent=2))

except Exception as e:
    print(f"Error: {str(e)}")
